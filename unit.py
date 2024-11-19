from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import openpyxl
from datetime import datetime
import os

def setup_browser():
    """
    Cấu hình trình duyệt và mở cửa sổ.
    """
    chrome_options = Options()
    chrome_options.add_experimental_option("detach", True)
    driver = webdriver.Chrome(options=chrome_options)
    return driver

def split_screen(driver1, driver2):
    """
    Chia đôi màn hình cho 2 trình duyệt.
    """
    screen_width = driver1.execute_script("return window.screen.width")
    screen_height = driver1.execute_script("return window.screen.height")

    # Cửa sổ trình duyệt 1
    driver1.set_window_position(0, 0)
    driver1.set_window_size(screen_width // 2, screen_height)

    # Cửa sổ trình duyệt 2
    driver2.set_window_position(screen_width // 2, 0)
    driver2.set_window_size(screen_width // 2, screen_height)

def setup_excel(template_path, template_sheet_name):
    """
    Chuẩn bị workbook và sheet từ template.
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"File template không tồn tại tại: {template_path}")

    workbook = openpyxl.load_workbook(template_path)
    if template_sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{template_sheet_name}' không tồn tại trong file.")

    template_sheet = workbook[template_sheet_name]
    now = datetime.now()
    new_sheet_name = f"So sanh {now.strftime('%d-%m %H-%M')}"
    new_sheet = workbook.copy_worksheet(template_sheet)
    new_sheet.title = new_sheet_name

    return workbook, new_sheet

def write_to_excel(sheet, start_row, start_column, data):
    """
    Ghi dữ liệu sản phẩm vào Excel từ dòng và cột bắt đầu.
    """
    for row_index, (name, price, url) in enumerate(data, start=start_row):
        sheet.cell(row=row_index, column=start_column, value=name)
        sheet.cell(row=row_index, column=start_column + 1, value=price)
        sheet.cell(row=row_index, column=start_column + 2, value=url)
