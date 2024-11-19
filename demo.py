from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from datetime import datetime
import os

# Giữ trình duyệt mở sau khi chạy
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

# Đường dẫn file Excel template
template_file = r'D:\Selenium\Rp So sanh.xlsx'

# Kiểm tra file template tồn tại
if not os.path.exists(template_file):
    raise FileNotFoundError(f"File template không tồn tại tại: {template_file}")

# Mở file Excel template
workbook = openpyxl.load_workbook(template_file)

# Lấy sheet template
template_sheet_name = 'template'
if template_sheet_name not in workbook.sheetnames:
    raise ValueError(f"Sheet '{template_sheet_name}' không tồn tại trong file.")
template_sheet = workbook[template_sheet_name]

# Tạo tên sheet mới với định dạng hợp lệ
now = datetime.now()
new_sheet_name = f"So sanh {now.strftime('%d-%m %H-%M')}"  # Thay : bằng -

# Tạo sheet mới từ template
new_sheet = workbook.copy_worksheet(template_sheet)
new_sheet.title = new_sheet_name

# Khởi tạo trình duyệt
driver = webdriver.Chrome(options=chrome_options)
driver2 = webdriver.Chrome(options=chrome_options)
screen_width = driver.execute_script("return window.screen.width")
screen_height = driver.execute_script("return window.screen.height")

# Mở trang Thegioididong trên driver chính
driver.get("https://www.thegioididong.com")
# Đặt cửa sổ của Thegioididong ở bên trái
driver.set_window_position(0, 0)  # Đặt vị trí ở góc trên bên trái
driver.set_window_size(screen_width // 2, screen_height)  # Đặt kích thước cửa sổ chiếm nửa màn hình bên trái


# Mở trang CellphoneS trên driver thứ hai
driver2.get("https://cellphones.com.vn")
# Đặt cửa sổ của CellphoneS ngay sát bên phải cửa sổ của Thegioididong
driver2.set_window_position(screen_width // 2, 0)  # Đặt vị trí ở nửa phải màn hình
driver2.set_window_size(screen_width // 2, screen_height)  # Đặt kích thước cửa sổ chiếm nửa màn hình bên phải

# Tìm kiếm sản phẩm trên Thegioididong
driver.find_element(By.ID, 'skw').send_keys('iphone 16 promax')
driver.find_element(By.XPATH, "//button[i[@class='icon-search']]").click()

# Sử dụng WebDriverWait để chờ các sản phẩm tải xong
WebDriverWait(driver, 10).until(
    EC.presence_of_all_elements_located((By.XPATH, "//li[contains(@class, 'item cat42')]"))
)

# Lấy tất cả các sản phẩm sau khi chúng đã tải xong
products = driver.find_elements(By.XPATH, "//li[contains(@class, 'item cat42')]")

# Bắt đầu ghi dữ liệu từ dòng thứ 3
start_row = 3  # Dòng bắt đầu ghi dữ liệu
start_column = 6  # Cột bắt đầu ghi dữ liệu (F)

for row_index, product in enumerate(products, start=start_row):
    try:
        # Lấy tên, giá và URL của sản phẩm
        product_name = product.find_element(By.XPATH, ".//h3").text.strip()
        product_price = product.find_element(By.XPATH, ".//strong[@class='price']").text.strip()
        product_url = product.find_element(By.XPATH, ".//a").get_attribute("href")

        # Ghi dữ liệu vào sheet, bắt đầu từ dòng 3
        new_sheet.cell(row=row_index, column=1, value=product_name)  # Cột 1: Tên sản phẩm
        new_sheet.cell(row=row_index, column=2, value=product_price)  # Cột 2: Giá tiền
        new_sheet.cell(row=row_index, column=3, value=product_url)   # Cột 3: URL sản phẩm

        print(f"Đã ghi: {product_name}, {product_price}, {product_url}")

    except Exception as e:
        print(f"Lỗi khi xử lý sản phẩm: {e}")

driver2.find_element(By.ID,'inp$earch').send_keys('iphone 16 promax')
driver2.find_element(By.XPATH, "//div[@class='input-group-btn']").click()

# Sử dụng WebDriverWait để chờ các sản phẩm tải xong
WebDriverWait(driver2, 10).until(
    EC.presence_of_all_elements_located((By.XPATH, '//div[@class="product-info-container product-item"]'))
)
# Lấy tất cả các sản phẩm sau khi chúng đã tải xong
products2 = driver2.find_elements(By.XPATH, '//div[@class="product-info-container product-item"]')

for row_index, product2 in enumerate(products2, start=start_row):
    try:
        # Lấy tên, giá và URL của sản phẩm
        product_name = product2.find_element(By.XPATH, ".//div[@class='product__name']/h3").text.strip()
        product_price = product2.find_element(By.XPATH, './/p[@class="product__price--show"]').text.strip()
        product_url = product2.find_element(By.XPATH, ".//a[@class='product__link button__link']").get_attribute("href")

        # Ghi dữ liệu vào sheet, bắt đầu từ cột 6 (F)
        new_sheet.cell(row=row_index, column=start_column, value=product_name)       # Cột 6 (F): Tên sản phẩm
        new_sheet.cell(row=row_index, column=start_column + 1, value=product_price)  # Cột 7 (G): Giá tiền
        new_sheet.cell(row=row_index, column=start_column + 2, value=product_url)    # Cột 8 (H): URL sản phẩm

        print(f"Đã ghi: {product_name}, {product_price}, {product_url}")

    except Exception as e:
        print(f"Lỗi khi xử lý sản phẩm: {e}")

# Đóng trình duyệt
driver.quit()
driver2.quit()

# Lưu workbook vào chính file template
workbook.save(template_file)
print(f"Đã lưu dữ liệu vào file Excel: {template_file}")
